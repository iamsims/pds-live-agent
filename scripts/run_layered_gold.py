"""Run the layered finder over data/gold_datasets.xlsx and log to xlsx.

For each query in the gold set:
  1. Router classifies → primary node.
  2. Worker for that node runs (Stage 1 directory walk + Stage 2 facets).
  3. We capture: router decision, candidates, each tool call (input + output
     preview), summary, elapsed, and whether any candidate matches the
     expected identifier(s).

Two output workbooks:
  - <out>.xlsx      one row per query
  - <out>_calls.xlsx  one row per tool call, joinable on `row_index`

Usage:

    cd /Users/skc/Desktop/Work/PDS_current_4.28/pydantic_code
    .venv/bin/python -m pydantic_code.scripts.run_layered_gold --limit 5 -v
    .venv/bin/python -m pydantic_code.scripts.run_layered_gold        # full set

Env:
    OPENAI_API_KEY     LLM credentials
    FAST_MCP_AUTH      bearer for both hosted MCP instances
"""

from __future__ import annotations

import argparse
import asyncio
import json
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pydantic_code.live_finder.pds_finder import LayeredFinder
from pydantic_code.run_eval import _extract_tool_calls

_REPO_ROOT = Path(__file__).resolve().parent.parent
_DEFAULT_GOLD = _REPO_ROOT / "data" / "gold_datasets.xlsx"
_DEFAULT_OUT_DIR = _REPO_ROOT / "output"
_CELL_MAX = 32_000  # well under Excel's 32767 hard limit


def _trunc(s: str, n: int = _CELL_MAX) -> str:
    if s is None:
        return ""
    s = str(s)
    return s if len(s) <= n else s[: n - 14] + "...(truncated)"


def load_gold_queries(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        paper, query, expected_raw = (row + (None, None, None))[:3]
        if not query:
            continue
        expected_tokens = [t.strip() for t in (expected_raw or "").split() if t.strip()]
        rows.append({
            "row_index": idx,
            "paper": paper or "",
            "query": query,
            "expected_ids_raw": expected_raw or "",
            "expected_ids": expected_tokens,
        })
    return rows


def check_match(
    candidates: list[Any],
    expected_ids: list[str],
) -> tuple[bool, list[str], list[int]]:
    """Case-insensitive substring match between expected ids and each
    candidate's dataset_id or path. Returns (matched, which_expected,
    which_candidate_indices).
    """
    if not expected_ids:
        return False, [], []
    matched_exp: list[str] = []
    matched_idx: list[int] = []
    for i, c in enumerate(candidates):
        haystacks: list[str] = []
        for f in ("dataset_id", "path"):
            v = getattr(c, f, None)
            if v:
                haystacks.append(str(v).lower())
        if not haystacks:
            continue
        for e in expected_ids:
            el = e.lower()
            if any(el in h or h in el for h in haystacks):
                if e not in matched_exp:
                    matched_exp.append(e)
                if i not in matched_idx:
                    matched_idx.append(i)
    return bool(matched_exp), matched_exp, matched_idx


async def run_all(
    queries: list[dict[str, Any]],
    *,
    model: str,
    effort: str,
    limit: int | None,
    start: int,
    verbose: bool,
) -> list[dict[str, Any]]:
    rows = queries[start : start + limit] if limit else queries[start:]
    total = len(rows)
    results: list[dict[str, Any]] = []

    async with LayeredFinder(model=model, reasoning_effort=effort) as lf:
        for i, row in enumerate(rows):
            t0 = time.monotonic()
            decision = output = None
            tool_calls: list[dict[str, Any]] = []
            error = None
            try:
                decision, output, messages = await lf.run_traced(row["query"])
                tool_calls = _extract_tool_calls(messages)
            except Exception as e:  # noqa: BLE001
                error = repr(e)
            dt = time.monotonic() - t0

            matched, matched_exp, matched_idx = (False, [], [])
            if output is not None:
                matched, matched_exp, matched_idx = check_match(
                    output.candidates, row["expected_ids"]
                )

            r = {
                **row,
                "router": decision,
                "output": output,
                "tool_calls": tool_calls,
                "matched": matched,
                "matched_expected": matched_exp,
                "matched_candidate_indices": matched_idx,
                "elapsed_s": round(dt, 2),
                "error": error,
            }
            results.append(r)

            if verbose:
                node = decision.primary_node if decision else "ERR"
                n_cands = len(output.candidates) if output else 0
                marker = "MATCH" if matched else ("ERR  " if error else "miss ")
                print(
                    f"[{i + 1:3d}/{total}] {marker}  node={node:<5}  "
                    f"calls={len(tool_calls):2d}  cands={n_cands:2d}  "
                    f"{dt:5.1f}s  | {row['query'][:70]}"
                )
    return results


def _row_for_results_sheet(r: dict[str, Any]) -> list[Any]:
    cand_ids: list[str] = []
    cand_paths: list[str] = []
    cand_reasonings: list[str] = []
    summary = ""
    if r.get("output"):
        for c in r["output"].candidates:
            cand_ids.append(c.dataset_id or "")
            cand_paths.append(c.path or "")
            cand_reasonings.append(c.reasoning or "")
        summary = r["output"].summary or ""

    tc_names = " → ".join(tc["tool_name"] for tc in r["tool_calls"])

    router = r.get("router")
    return [
        r["row_index"],
        _trunc(r["paper"], 400),
        _trunc(r["query"]),
        _trunc(r["expected_ids_raw"], 400),
        router.primary_node if router else "",
        router.secondary_node if router else "",
        router.confidence if router else "",
        _trunc(router.reasoning if router else "", 1000),
        bool(r["matched"]),
        "; ".join(r["matched_expected"]),
        "; ".join(str(i) for i in r["matched_candidate_indices"]),
        len(cand_ids),
        _trunc("\n".join(cand_ids)),
        _trunc("\n".join(cand_paths)),
        _trunc("\n---\n".join(cand_reasonings)),
        _trunc(summary, 4000),
        len(r["tool_calls"]),
        _trunc(tc_names, 4000),
        r["elapsed_s"],
        _trunc(r["error"] or "", 4000),
    ]


_RESULTS_HEADERS = [
    "row_index", "paper", "query", "expected_ids_raw",
    "router_node", "router_secondary", "router_confidence", "router_reasoning",
    "matched", "matched_expected_ids", "matched_candidate_indices",
    "candidate_count", "candidate_dataset_ids", "candidate_paths",
    "candidate_reasonings", "summary",
    "tool_call_count", "tool_call_sequence",
    "elapsed_s", "error",
]

_RESULTS_WIDTHS = {
    1: 8, 2: 30, 3: 50, 4: 30, 5: 10, 6: 12, 7: 12, 8: 40,
    9: 9, 10: 25, 11: 18, 12: 8, 13: 36, 14: 36, 15: 60, 16: 50,
    17: 8, 18: 60, 19: 10, 20: 40,
}


def _row_for_calls_sheet(r: dict[str, Any]) -> list[list[Any]]:
    out: list[list[Any]] = []
    for i, tc in enumerate(r["tool_calls"]):
        out.append([
            r["row_index"],
            i + 1,
            tc["tool_name"],
            _trunc(json.dumps(tc.get("tool_input") or {}, ensure_ascii=False), 8000),
            _trunc(tc.get("tool_output") or "", 16000),
        ])
    return out


def write_workbooks(out_path: Path, results: list[dict[str, Any]]) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # === main results workbook ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(_RESULTS_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in results:
        ws.append(_row_for_results_sheet(r))

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    match_col = _RESULTS_HEADERS.index("matched") + 1
    for row_idx, r in enumerate(results, start=2):
        cell = ws.cell(row=row_idx, column=match_col)
        if r["matched"]:
            cell.fill = green
        elif r["error"]:
            cell.fill = red

    for col, w in _RESULTS_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # Tool calls tab — inline so it's a single file
    cw = wb.create_sheet("Tool Calls")
    cw.append(["row_index", "call_index", "tool_name", "tool_input", "tool_output"])
    for cell in cw[1]:
        cell.font = Font(bold=True)
    for r in results:
        for row in _row_for_calls_sheet(r):
            cw.append(row)
    for col, w in {1: 9, 2: 10, 3: 24, 4: 50, 5: 80}.items():
        cw.column_dimensions[get_column_letter(col)].width = w
    cw.freeze_panes = "A2"

    # Summary
    s = wb.create_sheet("Summary")
    total = len(results)
    matched = sum(1 for r in results if r["matched"])
    errored = sum(1 for r in results if r["error"])
    total_calls = sum(len(r["tool_calls"]) for r in results)
    s.append(["metric", "value"])
    s.append(["total_queries", total])
    s.append(["matched", matched])
    s.append(["match_rate", round(matched / total, 3) if total else 0])
    s.append(["errored", errored])
    s.append(["total_tool_calls", total_calls])
    s.append(["avg_tool_calls_per_query",
              round(total_calls / total, 2) if total else 0])
    s.append(["total_elapsed_s", round(sum(r["elapsed_s"] for r in results), 1)])
    s.append([])

    nodes = Counter(r["router"].primary_node for r in results if r.get("router"))
    matches_per_node = Counter(
        r["router"].primary_node for r in results
        if r.get("router") and r["matched"]
    )
    s.append(["node", "total", "matched", "match_rate"])
    for node, count in nodes.most_common():
        m = matches_per_node.get(node, 0)
        s.append([node, count, m, round(m / count, 3) if count else 0])
    for cell in s[1]:
        cell.font = Font(bold=True)
    s.column_dimensions["A"].width = 28
    s.column_dimensions["B"].width = 14
    s.column_dimensions["C"].width = 10
    s.column_dimensions["D"].width = 12

    wb.save(out_path)
    return out_path


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--gold", type=Path, default=_DEFAULT_GOLD)
    p.add_argument("--out", type=Path, default=None,
                   help="Output xlsx path. Defaults to output/layered_gold_<ts>.xlsx")
    p.add_argument("--limit", type=int, default=None, help="Run only the first N queries.")
    p.add_argument("--start", type=int, default=0, help="Skip the first N queries (resume).")
    p.add_argument("--model", default="openai:gpt-5.2")
    p.add_argument("--effort", default="high", choices=["low", "medium", "high"])
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args()

    queries = load_gold_queries(args.gold)
    print(f"Loaded {len(queries)} queries from {args.gold}")
    if args.start:
        print(f"  skipping first {args.start}")
    if args.limit:
        print(f"  running {args.limit} queries")
    print()

    out = args.out or _DEFAULT_OUT_DIR / (
        f"layered_gold_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    t0 = time.monotonic()
    results = asyncio.run(
        run_all(queries, model=args.model, effort=args.effort,
                limit=args.limit, start=args.start, verbose=args.verbose)
    )
    elapsed = time.monotonic() - t0

    write_workbooks(out, results)

    matched = sum(1 for r in results if r["matched"])
    errored = sum(1 for r in results if r["error"])
    print()
    print(f"=== Done in {elapsed / 60:.1f}m ===")
    print(f"Wrote: {out}")
    print(f"  total:   {len(results)}")
    print(f"  matched: {matched} ({matched / len(results) * 100:.1f}%)" if results else "  matched: 0")
    print(f"  errored: {errored}")


if __name__ == "__main__":
    main()
