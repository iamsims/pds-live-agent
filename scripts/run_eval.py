"""Run layered and/or catalog evals over data/gold_datasets.xlsx and produce
a side-by-side comparison.

Layout:

    output/eval_<ts>/
        layered/
            results.jsonl, results.xlsx, summary.json, traces/
        catalog/
            results.jsonl, results.xlsx, summary.json, traces/
        comparison.json   # joined per-query metrics + aggregate delta

Both modes share:
  * --limit, --start    same query subset
  * --concurrency       same in-flight semaphore
  * --model, --effort   same model + reasoning effort

Usage:
    cd /Users/skc/Desktop/Work/PDS_current_4.28
    pydantic_code/.venv/bin/python -m pydantic_code.scripts.run_eval --limit 10
    pydantic_code/.venv/bin/python -m pydantic_code.scripts.run_eval --only catalog --limit 3
    pydantic_code/.venv/bin/python -m pydantic_code.scripts.run_eval --only layered -v

Env:
    OPENAI_API_KEY     LLM credentials
    FAST_MCP_AUTH      bearer for the hosted MCP instances
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

_PARENT_OF_PACKAGE = Path(__file__).resolve().parent.parent.parent
if str(_PARENT_OF_PACKAGE) not in sys.path:
    sys.path.insert(0, str(_PARENT_OF_PACKAGE))

import openpyxl  # noqa: E402
from dotenv import load_dotenv  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

load_dotenv()

from pydantic_ai.usage import RunUsage  # noqa: E402
from pydantic_code.finder import FindDatasetOutput, build_finder  # noqa: E402
from pydantic_code.live_finder.pds_finder import LiveFinder  # noqa: E402
from pydantic_code.scripts.eval_helpers import (  # noqa: E402
    check_match,
    extract_tool_calls,
    format_candidates,
    is_error_trace,
    serialize,
    sum_usage,
    usage_to_dict,
)

_REPO_ROOT = Path(__file__).resolve().parent.parent
_DEFAULT_GOLD = _REPO_ROOT / "data" / "gold_datasets.xlsx"
_OUT_BASE = _REPO_ROOT / "output"
_CELL_MAX = 32_000


def _trunc(s: str, n: int = _CELL_MAX) -> str:
    if s is None:
        return ""
    s = str(s)
    return s if len(s) <= n else s[: n - 14] + "...(truncated)"


# ---------------------------------------------------------------------------
# Gold dataset loader
# ---------------------------------------------------------------------------


def load_gold_queries(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        paper, query, expected_raw = (row + (None, None, None))[:3]
        if not query:
            continue
        expected_tokens = [t.strip() for t in (expected_raw or "").split() if t.strip()]
        rows.append({
            "row_index": idx,
            "paper": paper or "",
            "query": query,
            "expected_ids_raw": expected_raw or "",
            "expected_ids": expected_tokens,
        })
    return rows


def _load_jsonl_results(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                pass
    return rows


# ---------------------------------------------------------------------------
# Layered runner
# ---------------------------------------------------------------------------


async def _process_layered(
    lf: LiveFinder,
    row: dict[str, Any],
    sem: asyncio.Semaphore,
    idx: int,
    total: int,
    verbose: bool,
    jsonl_path: Path | None = None,
    jsonl_lock: asyncio.Lock | None = None,
    traces_dir: Path | None = None,
) -> dict[str, Any]:
    async with sem:
        print(f"[{idx + 1:3d}/{total}] start: {str(row['query'])[:80]}", flush=True)
        t0 = time.monotonic()
        decision = output = None
        tool_calls: list[dict[str, Any]] = []
        messages: list = []
        router_usage = RunUsage()
        worker_usage = RunUsage()
        error = None
        try:
            decision, output, messages, _ = await lf.run_traced(
                row["query"],
                router_usage=router_usage,
                worker_usage=worker_usage,
            )
            tool_calls = extract_tool_calls(messages)
        except Exception as e:  # noqa: BLE001
            error = repr(e)
            messages = [{"error": error}]
        usage = sum_usage(
            usage_to_dict(router_usage),
            usage_to_dict(worker_usage),
        )
        dt = time.monotonic() - t0

        matched, matched_exp, matched_idx = (False, [], [])
        if output is not None:
            matched, matched_exp, matched_idx = check_match(
                output.candidates, row["expected_ids"]
            )

        r: dict[str, Any] = {
            **row,
            "router_node": decision.primary_node if decision else "",
            "router_secondary": decision.secondary_node if decision else "",
            "router_confidence": decision.confidence if decision else 0,
            "router_reasoning": decision.reasoning if decision else "",
            "candidates": [c.model_dump() for c in output.candidates] if output else [],
            "summary": output.summary if output else "",
            "tool_calls": tool_calls,
            "usage": usage,
            "matched": matched,
            "matched_expected": matched_exp,
            "matched_candidate_indices": matched_idx,
            "elapsed_s": round(dt, 2),
            "error": error,
        }

        if verbose:
            node = r["router_node"] or "ERR"
            n_cands = len(r["candidates"])
            marker = "MATCH" if matched else ("ERR  " if error else "miss ")
            print(
                f"[{idx + 1:3d}/{total}] {marker}  node={node:<5}  "
                f"calls={len(tool_calls):2d}  cands={n_cands:2d}  "
                f"toks={usage['total_tokens']:>6}  {dt:5.1f}s  "
                f"| {row['query'][:60]}"
            )

        if jsonl_path is not None:
            line = json.dumps(r, default=str, ensure_ascii=False)
            if jsonl_lock is not None:
                async with jsonl_lock:
                    with jsonl_path.open("a", encoding="utf-8") as f:
                        f.write(line + "\n")
            else:
                with jsonl_path.open("a", encoding="utf-8") as f:
                    f.write(line + "\n")

        if traces_dir is not None:
            trace_data = {
                "row_index": row["row_index"],
                "query": row["query"],
                "row": row,
                "elapsed_s": r["elapsed_s"],
                "router_node": r["router_node"],
                "n_candidates": len(r["candidates"]),
                "n_tool_calls": len(tool_calls),
                "usage": r["usage"],
                "tool_calls": tool_calls,
                "matched": r["matched"],
                "matched_expected": r["matched_expected"],
                "full_trace": (
                    [serialize(m) for m in messages]
                    if not is_error_trace(messages)
                    else messages
                ),
                "error": error,
            }
            (traces_dir / f"r{row['row_index']:04d}_trace.json").write_text(
                json.dumps(trace_data, indent=2, default=str),
                encoding="utf-8",
            )
        return r


async def run_layered(
    queries: list[dict[str, Any]],
    *,
    model: str,
    effort: str,
    limit: int | None,
    start: int,
    concurrency: int,
    verbose: bool,
    jsonl_path: Path | None = None,
    traces_dir: Path | None = None,
) -> tuple[list[dict[str, Any]], float, float]:
    """Run layered eval. Returns ``(results, warm_elapsed_s, run_elapsed_s)``."""
    rows = queries[start : start + limit] if limit else queries[start:]

    resumed_results: list[dict[str, Any]] = []
    if jsonl_path is not None:
        resumed_results = _load_jsonl_results(jsonl_path)
        done_queries = {str(r.get("query", "")).strip() for r in resumed_results}
        done_queries.discard("")
        if done_queries:
            print(f"  resume: {len(done_queries)} query(s) already done; skipping")
        rows = [r for r in rows if str(r["query"]).strip() not in done_queries]

    total = len(rows)
    if total == 0:
        print("  nothing to run — all queries already complete in JSONL.")
        return sorted(resumed_results, key=lambda r: r["row_index"]), 0.0, 0.0

    sem = asyncio.Semaphore(max(1, concurrency))
    jsonl_lock = asyncio.Lock() if jsonl_path is not None else None

    async with LiveFinder(model=model, reasoning_effort=effort) as lf:
        t_warm = time.monotonic()
        await lf.warm()
        warm_elapsed = time.monotonic() - t_warm
        print(f"  pre-warmed {len(lf._workers)} worker(s) in {warm_elapsed:.1f}s")

        t_run = time.monotonic()
        raw = await asyncio.gather(
            *(
                _process_layered(
                    lf, row, sem, i, total, verbose,
                    jsonl_path=jsonl_path, jsonl_lock=jsonl_lock,
                    traces_dir=traces_dir,
                )
                for i, row in enumerate(rows)
            )
        )
        run_elapsed = time.monotonic() - t_run

    all_results = resumed_results + list(raw)
    return sorted(all_results, key=lambda r: r["row_index"]), warm_elapsed, run_elapsed


def _build_layered_summary(
    results: list[dict[str, Any]],
    *,
    warm_elapsed_s: float,
    run_elapsed_s: float,
) -> dict[str, Any]:
    n = len(results)
    matched = sum(1 for r in results if r["matched"])
    errored = sum(1 for r in results if r["error"])
    total_calls = sum(len(r["tool_calls"]) for r in results)
    tot_in = sum((r.get("usage") or {}).get("input_tokens", 0) for r in results)
    tot_out = sum((r.get("usage") or {}).get("output_tokens", 0) for r in results)
    tot_reqs = sum((r.get("usage") or {}).get("requests", 0) for r in results)
    sum_elapsed = sum(r["elapsed_s"] for r in results)

    tool_hist: Counter[str] = Counter()
    for r in results:
        for tc in r["tool_calls"]:
            tool_hist[tc["tool_name"]] += 1

    nodes = Counter(r["router_node"] for r in results if r.get("router_node"))
    matches_per_node = Counter(
        r["router_node"] for r in results
        if r.get("router_node") and r["matched"]
    )
    per_node = [
        {
            "node": node,
            "n": count,
            "matched": matches_per_node.get(node, 0),
            "match_rate": round(matches_per_node.get(node, 0) / count, 3) if count else 0,
        }
        for node, count in nodes.most_common()
    ]

    return {
        "mode": "layered",
        "n_queries": n,
        "n_matched": matched,
        "match_rate": round(matched / n, 3) if n else 0,
        "n_errored": errored,
        "warm_elapsed_s": round(warm_elapsed_s, 2),
        "run_elapsed_s": round(run_elapsed_s, 2),
        "total_query_elapsed_s": round(sum_elapsed, 1),
        "avg_query_elapsed_s": round(sum_elapsed / n, 2) if n else 0,
        "total_tool_calls": total_calls,
        "avg_tool_calls": round(total_calls / n, 2) if n else 0,
        "total_input_tokens": tot_in,
        "total_output_tokens": tot_out,
        "total_tokens": tot_in + tot_out,
        "avg_total_tokens": round((tot_in + tot_out) / n, 0) if n else 0,
        "total_llm_requests": tot_reqs,
        "tool_call_histogram": dict(tool_hist.most_common()),
        "per_node": per_node,
    }


def _write_layered_xlsx(out_path: Path, results: list[dict[str, Any]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "row_index", "paper", "query", "expected_ids_raw",
        "router_node", "router_secondary", "router_confidence", "router_reasoning",
        "matched", "matched_expected_ids", "matched_candidate_indices",
        "candidate_count", "candidate_dataset_ids", "candidate_paths",
        "candidate_reasonings", "summary",
        "tool_call_count", "tool_call_sequence",
        "elapsed_s",
        "input_tokens", "output_tokens", "total_tokens", "requests",
        "error",
    ]
    widths = {
        1: 8, 2: 30, 3: 50, 4: 30, 5: 10, 6: 12, 7: 12, 8: 40,
        9: 9, 10: 25, 11: 18, 12: 8, 13: 36, 14: 36, 15: 60, 16: 50,
        17: 8, 18: 60, 19: 10,
        20: 12, 21: 12, 22: 12, 23: 10,
        24: 40,
    }

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in results:
        cand_ids = [c.get("dataset_id", "") for c in (r.get("candidates") or [])]
        cand_paths = [c.get("path", "") for c in (r.get("candidates") or [])]
        cand_reasonings = [c.get("reasoning", "") for c in (r.get("candidates") or [])]
        usage = r.get("usage") or {}
        tc_names = " → ".join(tc["tool_name"] for tc in r["tool_calls"])
        ws.append([
            r["row_index"],
            _trunc(r["paper"], 400),
            _trunc(r["query"]),
            _trunc(r["expected_ids_raw"], 400),
            r.get("router_node", ""),
            r.get("router_secondary", ""),
            r.get("router_confidence", ""),
            _trunc(r.get("router_reasoning", "") or "", 1000),
            bool(r["matched"]),
            "; ".join(r["matched_expected"]),
            "; ".join(str(i) for i in r["matched_candidate_indices"]),
            len(cand_ids),
            _trunc("\n".join(cand_ids)),
            _trunc("\n".join(cand_paths)),
            _trunc("\n---\n".join(cand_reasonings)),
            _trunc(r.get("summary", ""), 4000),
            len(r["tool_calls"]),
            _trunc(tc_names, 4000),
            r["elapsed_s"],
            usage.get("input_tokens", 0),
            usage.get("output_tokens", 0),
            usage.get("total_tokens", 0),
            usage.get("requests", 0),
            _trunc(r["error"] or "", 4000),
        ])

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    match_col = headers.index("matched") + 1
    for row_idx, r in enumerate(results, start=2):
        cell = ws.cell(row=row_idx, column=match_col)
        if r["matched"]:
            cell.fill = green
        elif r["error"]:
            cell.fill = red

    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    cw = wb.create_sheet("Tool Calls")
    cw.append(["row_index", "call_index", "tool_name", "tool_input", "tool_output"])
    for cell in cw[1]:
        cell.font = Font(bold=True)
    for r in results:
        for i, tc in enumerate(r["tool_calls"]):
            cw.append([
                r["row_index"],
                i + 1,
                tc["tool_name"],
                _trunc(json.dumps(tc.get("tool_input") or {}, ensure_ascii=False), 8000),
                _trunc(tc.get("tool_output") or "", 16000),
            ])
    for col, w in {1: 9, 2: 10, 3: 24, 4: 50, 5: 80}.items():
        cw.column_dimensions[get_column_letter(col)].width = w
    cw.freeze_panes = "A2"

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Catalog runner
# ---------------------------------------------------------------------------


async def _warm_mcp(agent) -> None:
    pings = []
    for ts in getattr(agent, "toolsets", ()) or ():
        list_fn = getattr(ts, "list_tools", None)
        if callable(list_fn):
            pings.append(list_fn())
    if pings:
        await asyncio.gather(*pings, return_exceptions=True)


async def _process_catalog(
    idx: int,
    row: dict,
    agent,
    traces_dir: Path,
    sem: asyncio.Semaphore,
    jsonl_path: Path | None = None,
    jsonl_lock: asyncio.Lock | None = None,
) -> dict:
    query = row["query"]
    async with sem:
        print(f"[{idx + 1}] start: {str(query)[:80]}", flush=True)
        started = time.monotonic()
        output: FindDatasetOutput | None = None
        usage = RunUsage()
        trace: list = []
        try:
            result = await agent.run(query, usage=usage)
            output = result.output
            trace = list(result.all_messages())
        except Exception as e:
            err = f"{type(e).__name__}: {e}"
            print(f"      → FAILED: {err}")
            trace = [{"error": err}]
        elapsed = time.monotonic() - started

        tool_calls = (
            extract_tool_calls(trace)
            if trace and not is_error_trace(trace)
            else []
        )
        tool_hist = Counter(tc["tool_name"] for tc in tool_calls)
        usage_d = usage_to_dict(usage)

        expected_raw = row.get("expected_ids_raw") or row.get("expected_ids") or ""
        expected_tokens = row.get("expected_ids") if isinstance(row.get("expected_ids"), list) else [
            t.strip() for t in str(expected_raw).split() if t.strip()
        ]
        matched, matched_exp, matched_idx = (False, [], [])
        if output is not None:
            matched, matched_exp, matched_idx = check_match(
                output.candidates, expected_tokens
            )

        n_cands = len(output.candidates) if output else 0
        marker = "MATCH" if matched else ("ERR  " if is_error_trace(trace) else "miss ")
        print(
            f"      → {marker}  {n_cands} cands, {len(tool_calls)} tool calls, "
            f"{usage_d.get('total_tokens', 0)} tokens, {elapsed:.1f}s"
        )

        trace_data = {
            "query": query,
            "row": row,
            "elapsed_s": round(elapsed, 2),
            "n_candidates": n_cands,
            "n_tool_calls": len(tool_calls),
            "tool_call_histogram": dict(tool_hist),
            "usage": usage_d,
            "tool_calls": tool_calls,
            "full_trace": (
                [serialize(m) for m in trace]
                if not is_error_trace(trace)
                else trace
            ),
        }
        trace_filename = (
            f"r{row['row_index']:04d}_trace.json"
            if row.get("row_index") is not None
            else f"{idx:03d}_trace.json"
        )
        (traces_dir / trace_filename).write_text(
            json.dumps(trace_data, indent=2, default=str)
        )

        result_row = {
            **row,
            "_idx": idx,
            "elapsed_s": round(elapsed, 2),
            "n_tool_calls": len(tool_calls),
            "tool_hist": dict(tool_hist),
            "tool_calls": tool_calls,
            "usage": usage_d,
            "n_candidates": n_cands,
            "matched": matched,
            "matched_expected": matched_exp,
            "matched_candidate_indices": matched_idx,
            "catalog_ids": "\n".join(
                c.dataset_id for c in (output.candidates if output else []) if c.dataset_id
            ),
            "catalog_candidates": format_candidates(output) if output else "ERROR",
            "catalog_summary": output.summary if output else "ERROR",
        }
        if jsonl_path is not None:
            line = json.dumps(result_row, default=str, ensure_ascii=False)
            if jsonl_lock is not None:
                async with jsonl_lock:
                    with jsonl_path.open("a", encoding="utf-8") as f:
                        f.write(line + "\n")
            else:
                with jsonl_path.open("a", encoding="utf-8") as f:
                    f.write(line + "\n")
        return result_row


async def run_catalog(
    queries: list[dict[str, Any]],
    *,
    model: str,
    effort: str,
    limit: int | None,
    concurrency: int,
    out_dir: Path,
) -> tuple[list[dict[str, Any]], float, float]:
    """Run catalog eval. Returns ``(results, warm_elapsed_s, run_elapsed_s)``."""
    rows = queries[:limit] if limit else queries

    out_dir.mkdir(parents=True, exist_ok=True)
    traces_dir = out_dir / "traces"
    traces_dir.mkdir(parents=True, exist_ok=True)
    jsonl_path = out_dir / "results.jsonl"

    resumed_results: list[dict[str, Any]] = []
    if jsonl_path.exists():
        resumed_results = _load_jsonl_results(jsonl_path)
        done_queries = {str(r.get("query", "")).strip() for r in resumed_results}
        done_queries.discard("")
        if done_queries:
            print(f"  resume: {len(done_queries)} query(s) already done; skipping")
        rows = [q for q in rows if str(q["query"]).strip() not in done_queries]

    if not rows:
        print("  nothing to run — all queries already complete in JSONL.")
        results = sorted(resumed_results, key=lambda r: r.get("row_index", 0))
        return results, 0.0, 0.0

    agent = build_finder(kind="catalog", model=model, reasoning_effort=effort)
    sem = asyncio.Semaphore(concurrency)
    jsonl_lock = asyncio.Lock()

    async with agent:
        t_warm = time.monotonic()
        await _warm_mcp(agent)
        warm_elapsed = time.monotonic() - t_warm
        print(f"  pre-warmed catalog MCP in {warm_elapsed:.1f}s")

        started = time.monotonic()
        raw = await asyncio.gather(
            *[
                _process_catalog(i, r, agent, traces_dir, sem,
                                 jsonl_path=jsonl_path, jsonl_lock=jsonl_lock)
                for i, r in enumerate(rows)
            ]
        )
        run_elapsed = time.monotonic() - started

    for r in raw:
        r.pop("_idx", None)
    results = sorted(resumed_results + list(raw), key=lambda r: r.get("row_index", 0))
    return results, warm_elapsed, run_elapsed


def _build_catalog_summary(
    results: list[dict[str, Any]],
    *,
    warm_elapsed_s: float,
    run_elapsed_s: float,
) -> dict[str, Any]:
    n = len(results)
    if not n:
        return {}
    total_tool_calls = sum(r["n_tool_calls"] for r in results)
    total_input = sum(r["usage"].get("input_tokens", 0) for r in results)
    total_output = sum(r["usage"].get("output_tokens", 0) for r in results)
    total_tokens = total_input + total_output
    total_requests = sum(r["usage"].get("requests", 0) for r in results)
    n_matched = sum(1 for r in results if r.get("matched"))
    n_errored = sum(1 for r in results if r.get("catalog_summary") == "ERROR")
    sum_query_elapsed = sum(r["elapsed_s"] for r in results)

    tool_hist = Counter()
    for r in results:
        tool_hist.update(r.get("tool_hist", {}))

    return {
        "mode": "catalog",
        "n_queries": n,
        "n_matched": n_matched,
        "match_rate": round(n_matched / n, 3),
        "n_errored": n_errored,
        "warm_elapsed_s": round(warm_elapsed_s, 2),
        "run_elapsed_s": round(run_elapsed_s, 2),
        "total_query_elapsed_s": round(sum_query_elapsed, 1),
        "avg_query_elapsed_s": round(sum_query_elapsed / n, 2),
        "total_tool_calls": total_tool_calls,
        "avg_tool_calls": round(total_tool_calls / n, 2),
        "total_input_tokens": total_input,
        "total_output_tokens": total_output,
        "total_tokens": total_tokens,
        "avg_total_tokens": round(total_tokens / n, 0),
        "total_llm_requests": total_requests,
        "tool_call_histogram": dict(tool_hist.most_common()),
    }


def _write_catalog_xlsx(path: Path, results: list[dict]) -> None:
    headers = [
        "Paper", "Query", "Expected Identifiers",
        "Matched", "Matched Expected IDs", "Matched Candidate Indices",
        "Catalog IDs", "Catalog Candidates", "Catalog Summary",
        "Elapsed (s)", "Tool Calls", "Tool Histogram",
        "Input Tokens", "Output Tokens", "Total Tokens", "Requests",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalog Eval"
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = header_font
    wrap = Alignment(wrap_text=True, vertical="top")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    for row_idx, r in enumerate(results, 2):
        u = r.get("usage", {})
        values = [
            r.get("paper"),
            r.get("query"),
            r.get("expected_ids_raw", r.get("expected_ids")),
            bool(r.get("matched")),
            "; ".join(r.get("matched_expected") or []),
            "; ".join(str(i) for i in (r.get("matched_candidate_indices") or [])),
            r.get("catalog_ids"),
            r.get("catalog_candidates"),
            r.get("catalog_summary"),
            r.get("elapsed_s"),
            r.get("n_tool_calls"),
            json.dumps(r.get("tool_hist", {})),
            u.get("input_tokens", 0),
            u.get("output_tokens", 0),
            u.get("total_tokens", 0),
            u.get("requests", 0),
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v).alignment = wrap
        match_col = headers.index("Matched") + 1
        cell = ws.cell(row=row_idx, column=match_col)
        if r.get("matched"):
            cell.fill = green
        elif r.get("catalog_summary") == "ERROR":
            cell.fill = red
    for col_idx, header in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        if "Candidates" in header or "Summary" in header:
            ws.column_dimensions[letter].width = 60
        elif "Histogram" in header:
            ws.column_dimensions[letter].width = 40
        elif "IDs" in header or "Identifiers" in header:
            ws.column_dimensions[letter].width = 40
        elif header == "Query":
            ws.column_dimensions[letter].width = 50
        else:
            ws.column_dimensions[letter].width = 16
    wb.save(path)


# ---------------------------------------------------------------------------
# Comparison printer
# ---------------------------------------------------------------------------


def _print_compare(layered: dict | None, catalog: dict | None) -> dict:
    keys = [
        ("Queries", "n_queries", "{}"),
        ("Matched", "n_matched", "{}"),
        ("Match rate", "match_rate", "{:.1%}"),
        ("Errored", "n_errored", "{}"),
        ("Warm (s)", "warm_elapsed_s", "{:.1f}"),
        ("Run wall-clock (s)", "run_elapsed_s", "{:.1f}"),
        ("Sum per-query (s)", "total_query_elapsed_s", "{:.1f}"),
        ("Avg per-query (s)", "avg_query_elapsed_s", "{:.2f}"),
        ("Tool calls", "total_tool_calls", "{}"),
        ("Avg tool calls", "avg_tool_calls", "{:.2f}"),
        ("Total tokens", "total_tokens", "{:,}"),
        ("  input", "total_input_tokens", "{:,}"),
        ("  output", "total_output_tokens", "{:,}"),
        ("Avg tokens", "avg_total_tokens", "{:.0f}"),
        ("LLM requests", "total_llm_requests", "{}"),
    ]
    print()
    print("=" * 72)
    print(f"{'Metric':<22}  {'Layered':>20}  {'Catalog':>20}")
    print("-" * 72)
    comp: dict[str, Any] = {}
    for label, key, fmt in keys:
        lv = (layered or {}).get(key)
        cv = (catalog or {}).get(key)
        ls = fmt.format(lv) if lv is not None else "—"
        cs = fmt.format(cv) if cv is not None else "—"
        print(f"{label:<22}  {ls:>20}  {cs:>20}")
        comp[key] = {"layered": lv, "catalog": cv}
    print("=" * 72)
    return comp


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--gold", type=Path, default=_DEFAULT_GOLD)
    p.add_argument("--out", type=Path, default=None,
                   help="Parent dir. Defaults to output/eval_<ts>/.")
    p.add_argument("--limit", type=int, default=None,
                   help="Run only the first N queries (same N for both modes).")
    p.add_argument("--start", type=int, default=0,
                   help="(Layered only) Skip the first N queries.")
    p.add_argument("--concurrency", type=int, default=3,
                   help="In-flight queries per mode (default: 3).")
    p.add_argument("--model", default="openai:gpt-5.2")
    p.add_argument("--effort", default="high", choices=["low", "medium", "high"])
    p.add_argument("--only", choices=["layered", "catalog"], default=None,
                   help="Run only one side instead of both.")
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    parent = args.out or (_OUT_BASE / f"eval_{ts}")
    parent.mkdir(parents=True, exist_ok=True)
    layered_dir = parent / "layered"
    catalog_dir = parent / "catalog"

    queries = load_gold_queries(args.gold)
    print(f"=== eval run @ {ts} ===")
    print(f"Parent: {parent}")
    print(f"Loaded {len(queries)} queries from {args.gold.name}")
    print(f"Flags: limit={args.limit} start={args.start} "
          f"concurrency={args.concurrency} model={args.model} effort={args.effort}")

    t0 = time.monotonic()

    async def _go() -> tuple[dict | None, dict | None]:
        layered_sum: dict | None = None
        catalog_sum: dict | None = None

        if args.only != "catalog":
            print(f"\n--- LAYERED ---")
            layered_dir.mkdir(parents=True, exist_ok=True)
            (layered_dir / "traces").mkdir(parents=True, exist_ok=True)
            results, warm_s, run_s = await run_layered(
                queries,
                model=args.model, effort=args.effort,
                limit=args.limit, start=args.start,
                concurrency=args.concurrency, verbose=args.verbose,
                jsonl_path=layered_dir / "results.jsonl",
                traces_dir=layered_dir / "traces",
            )
            _write_layered_xlsx(layered_dir / "results.xlsx", results)
            layered_sum = _build_layered_summary(results, warm_elapsed_s=warm_s, run_elapsed_s=run_s)
            (layered_dir / "summary.json").write_text(json.dumps(layered_sum, indent=2))

        if args.only != "layered":
            print(f"\n--- CATALOG ---")
            results, warm_s, run_s = await run_catalog(
                queries,
                model=args.model, effort=args.effort,
                limit=args.limit, concurrency=args.concurrency,
                out_dir=catalog_dir,
            )
            _write_catalog_xlsx(catalog_dir / "results.xlsx", results)
            catalog_sum = _build_catalog_summary(results, warm_elapsed_s=warm_s, run_elapsed_s=run_s)
            (catalog_dir / "summary.json").write_text(json.dumps(catalog_sum, indent=2))

        return layered_sum, catalog_sum

    layered_sum, catalog_sum = asyncio.run(_go())
    total = time.monotonic() - t0

    comp = _print_compare(layered_sum, catalog_sum)
    (parent / "comparison.json").write_text(json.dumps({
        "ts": ts,
        "flags": {
            "limit": args.limit, "start": args.start,
            "concurrency": args.concurrency,
            "model": args.model, "effort": args.effort,
        },
        "total_wallclock_s": round(total, 2),
        "metrics": comp,
        "layered_summary": layered_sum,
        "catalog_summary": catalog_sum,
    }, indent=2))
    print(f"\nWrote: {parent}")
    print(f"Total wall-clock: {total:.1f}s ({total / 60:.1f}m)")


if __name__ == "__main__":
    main()
