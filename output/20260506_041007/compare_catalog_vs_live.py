"""
Compare Catalog vs Live mode performance for PDS eval run 20260430_181930.
Generates side-by-side graphs from trace files and results.xlsx.
"""
import json
import os
import re
from datetime import datetime, timezone
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

BASE = os.path.dirname(os.path.abspath(__file__))
CATALOG_DIR = os.path.join(BASE, "traces", "catalog")
LIVE_DIR = os.path.join(BASE, "traces", "live")
RESULTS_FILE = os.path.join(BASE, "results.xlsx")
OUT_DIR = os.path.join(BASE, "graphs")
os.makedirs(OUT_DIR, exist_ok=True)

# ── helpers ──────────────────────────────────────────────────────────────────
def load_traces(trace_dir):
    """Return list of dicts with per-query metrics, sorted by file index."""
    entries = []
    for fname in sorted(os.listdir(trace_dir)):
        if not fname.endswith("_trace.json"):
            continue
        idx = int(fname.split("_")[0])
        with open(os.path.join(trace_dir, fname)) as f:
            data = json.load(f)

        tool_calls = len(data.get("tool_calls", []))
        full_trace = data.get("full_trace", [])
        trace_steps = len(full_trace)

        # Execution time: difference between first and last timestamp
        exec_time = 0.0
        timestamps = []
        for step in full_trace:
            if not isinstance(step, dict):
                continue
            ts = step.get("timestamp")
            if ts:
                timestamps.append(ts)
        if len(timestamps) >= 2:
            t0 = datetime.fromisoformat(str(timestamps[0]))
            t1 = datetime.fromisoformat(str(timestamps[-1]))
            exec_time = (t1 - t0).total_seconds()

        # Tool names used
        tool_names = [tc.get("tool_name", "") for tc in data.get("tool_calls", [])]

        query = data.get("query", "")
        paper = data.get("row", {}).get("paper_short", "")

        entries.append({
            "idx": idx,
            "query": query,
            "paper": paper,
            "tool_calls": tool_calls,
            "trace_steps": trace_steps,
            "exec_time": exec_time,
            "tool_names": tool_names,
        })
    return entries


def load_results():
    """Return per-row dicts from results.xlsx with dataset counts and expected-hit info."""
    wb = openpyxl.load_workbook(RESULTS_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, r))
        rows.append(d)
    return rows


def count_ids(id_string):
    """Count newline-separated dataset IDs."""
    if not id_string or str(id_string).strip() == "":
        return 0
    return len([x for x in str(id_string).strip().split("\n") if x.strip()])


# ── ID normalization (version suffix + slash/underscore only) ─────────────────


def normalize_pds_id(raw_id):
    """Normalize for comparison: strip ::version suffixes and unify / vs _ in PDS3 IDs."""
    s = raw_id.strip()
    # Remove trailing version like ::1.0, ::1.1, ::3.0
    s = re.sub(r'::[\d.]+$', '', s)
    s = s.upper()
    # Normalise slash vs underscore in PDS3 target-list segments
    # e.g. MESS-E/V/H/SW- → MESS-E_V_H_SW-  and  CO-E/SW/J/S- → CO-E_SW_J_S-
    s = re.sub(r'(?<=-)[A-Z0-9]+(/[A-Z0-9]+)+(?=-)', lambda m: m.group().replace('/', '_'), s)
    return s


def match_expected_ids(expected_str, found_str):
    """Check if any expected ID appears in the found IDs.

    Returns (hit: bool, rationale: str) explaining the match or why it failed.

    Only normalizes:
      1. Version suffix: ::1.0 etc. stripped from both sides
      2. Slash/underscore: E/V/H/SW treated same as E_V_H_SW in PDS3 IDs
      3. PDS3 dataset version difference: -V1.0 vs -V2.0
    No bundle↔collection or PDS3↔PDS4 equivalence mapping.
    """
    if not expected_str or not found_str:
        return False, "No expected or found IDs"

    expected_ids = [x.strip() for x in str(expected_str).split("\n") if x.strip()]
    found_ids = [x.strip() for x in str(found_str).split("\n") if x.strip()]

    # Build dict: normalized → original found ID
    found_norm_to_raw = {}
    for fid in found_ids:
        found_norm_to_raw[normalize_pds_id(fid)] = fid

    matches = []  # collect all per-expected-ID results

    for eid in expected_ids:
        norm_eid = normalize_pds_id(eid)
        raw_eid = eid.strip()

        # 1) Exact raw match
        if raw_eid in [f.strip() for f in found_ids]:
            matches.append((True, f"Exact match: {raw_eid}"))
            continue

        # 2) Normalized match (strips ::version, unifies / vs _)
        if norm_eid in found_norm_to_raw:
            matched_raw = found_norm_to_raw[norm_eid]
            if matched_raw.strip() == raw_eid:
                matches.append((True, f"Exact match: {raw_eid}"))
            else:
                reasons = []
                if re.search(r'::[\d.]+$', raw_eid) or re.search(r'::[\d.]+$', matched_raw):
                    reasons.append("version suffix stripped")
                if '/' in raw_eid and '_' in matched_raw or '_' in raw_eid and '/' in matched_raw:
                    reasons.append("slash/underscore normalized")
                if not reasons:
                    reasons.append("normalized equivalent")
                matches.append((True, f"{' + '.join(reasons)}: {raw_eid} ≈ {matched_raw}"))
            continue

        # 3) PDS3 dataset version difference (e.g., -V1.0 vs -V3.0)
        base_no_version = re.sub(r'-V[\d.]+$', '', norm_eid)
        version_match_found = None
        if base_no_version != norm_eid:  # only if there was a version to strip
            for fn, fn_raw in found_norm_to_raw.items():
                fn_no_version = re.sub(r'-V[\d.]+$', '', fn)
                if base_no_version == fn_no_version:
                    version_match_found = fn_raw
                    break
        if version_match_found:
            matches.append((True, f"Different version of same dataset: {raw_eid} ≈ {version_match_found}"))
            continue

        matches.append((False, f"No match for {raw_eid}"))

    any_hit = any(m[0] for m in matches)
    rationale = "; ".join(m[1] for m in matches)
    return any_hit, rationale


def check_expected_hit(expected_str, found_str):
    """Backward-compatible wrapper that returns just the bool."""
    hit, _ = match_expected_ids(expected_str, found_str)
    return hit


# ── load data ────────────────────────────────────────────────────────────────
catalog_traces = load_traces(CATALOG_DIR)
live_traces = load_traces(LIVE_DIR)
results = load_results()

n = len(catalog_traces)
indices = np.arange(n)
labels = [f"Q{i}" for i in range(n)]
short_labels = []
for i, row in enumerate(results):
    paper = row.get("Paper Short", "")
    short_labels.append(f"Q{i}\n{paper}")

cat_tool_calls = [t["tool_calls"] for t in catalog_traces]
live_tool_calls = [t["tool_calls"] for t in live_traces]
cat_steps = [t["trace_steps"] for t in catalog_traces]
live_steps = [t["trace_steps"] for t in live_traces]
cat_time = [t["exec_time"] for t in catalog_traces]
live_time = [t["exec_time"] for t in live_traces]

# Dataset counts from results
cat_dataset_counts = [count_ids(r.get("Catalog IDs")) for r in results]
live_dataset_counts = [count_ids(r.get("Live IDs")) for r in results]

# Check for catalog errors (rows where Catalog IDs is None or "ERROR")
cat_has_error = []
live_has_error = []
for r in results:
    cat_ids = r.get("Catalog IDs")
    live_ids = r.get("Live IDs")
    cat_has_error.append(cat_ids is None or str(cat_ids).strip() == "" or str(cat_ids).strip() == "ERROR")
    live_has_error.append(live_ids is None or str(live_ids).strip() == "" or str(live_ids).strip() == "ERROR")

# Expected hit check
cat_hit = [check_expected_hit(r.get("Expected Identifiers"), r.get("Catalog IDs")) for r in results]
live_hit = [check_expected_hit(r.get("Expected Identifiers"), r.get("Live IDs")) for r in results]

# ── styling ──────────────────────────────────────────────────────────────────
plt.rcParams.update({
    "figure.facecolor": "white",
    "axes.facecolor": "#f8f8f8",
    "axes.grid": True,
    "grid.alpha": 0.3,
    "font.size": 11,
})
CAT_COLOR = "#2196F3"   # blue
LIVE_COLOR = "#FF9800"  # orange
CAT_COLOR2 = "#1565C0"
LIVE_COLOR2 = "#E65100"
bar_width = 0.35

# ── GRAPH 1: Tool Calls per Query (grouped bar) ─────────────────────────────
fig, ax = plt.subplots(figsize=(14, 6))
ax.bar(indices - bar_width/2, cat_tool_calls, bar_width, label="Catalog", color=CAT_COLOR, edgecolor="white")
ax.bar(indices + bar_width/2, live_tool_calls, bar_width, label="Live", color=LIVE_COLOR, edgecolor="white")
ax.set_xlabel("Query")
ax.set_ylabel("Number of Tool Calls")
ax.set_title("Tool Calls per Query: Catalog vs Live")
ax.set_xticks(indices)
ax.set_xticklabels(labels)
ax.legend()
# Add averages as horizontal lines
ax.axhline(np.mean(cat_tool_calls), color=CAT_COLOR2, linestyle="--", alpha=0.7, label=f"Catalog avg: {np.mean(cat_tool_calls):.1f}")
ax.axhline(np.mean(live_tool_calls), color=LIVE_COLOR2, linestyle="--", alpha=0.7, label=f"Live avg: {np.mean(live_tool_calls):.1f}")
ax.legend()
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "01_tool_calls_per_query.png"), dpi=150)
plt.close(fig)

# ── GRAPH 2: Execution Time per Query (grouped bar) ─────────────────────────
fig, ax = plt.subplots(figsize=(14, 6))
ax.bar(indices - bar_width/2, [t/60 for t in cat_time], bar_width, label="Catalog", color=CAT_COLOR, edgecolor="white")
ax.bar(indices + bar_width/2, [t/60 for t in live_time], bar_width, label="Live", color=LIVE_COLOR, edgecolor="white")
ax.set_xlabel("Query")
ax.set_ylabel("Execution Time (minutes)")
ax.set_title("Execution Time per Query: Catalog vs Live")
ax.set_xticks(indices)
ax.set_xticklabels(labels)
ax.axhline(np.mean(cat_time)/60, color=CAT_COLOR2, linestyle="--", alpha=0.7)
ax.axhline(np.mean(live_time)/60, color=LIVE_COLOR2, linestyle="--", alpha=0.7)
ax.legend()
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "02_exec_time_per_query.png"), dpi=150)
plt.close(fig)

# ── GRAPH 3: Trace Steps per Query (grouped bar) ────────────────────────────
fig, ax = plt.subplots(figsize=(14, 6))
ax.bar(indices - bar_width/2, cat_steps, bar_width, label="Catalog", color=CAT_COLOR, edgecolor="white")
ax.bar(indices + bar_width/2, live_steps, bar_width, label="Live", color=LIVE_COLOR, edgecolor="white")
ax.set_xlabel("Query")
ax.set_ylabel("Number of Trace Steps (LLM rounds)")
ax.set_title("Trace Steps per Query: Catalog vs Live")
ax.set_xticks(indices)
ax.set_xticklabels(labels)
ax.axhline(np.mean(cat_steps), color=CAT_COLOR2, linestyle="--", alpha=0.7)
ax.axhline(np.mean(live_steps), color=LIVE_COLOR2, linestyle="--", alpha=0.7)
ax.legend()
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "03_trace_steps_per_query.png"), dpi=150)
plt.close(fig)

# ── GRAPH 4: Datasets Found per Query (grouped bar) ─────────────────────────
fig, ax = plt.subplots(figsize=(14, 6))
cat_bars = ax.bar(indices - bar_width/2, cat_dataset_counts, bar_width, label="Catalog", color=CAT_COLOR, edgecolor="white")
live_bars = ax.bar(indices + bar_width/2, live_dataset_counts, bar_width, label="Live", color=LIVE_COLOR, edgecolor="white")
# Mark errors with red X
for i in range(n):
    if cat_has_error[i]:
        ax.text(i - bar_width/2, 0.3, "ERR", ha="center", va="bottom", color="red", fontweight="bold", fontsize=8)
    if live_has_error[i]:
        ax.text(i + bar_width/2, 0.3, "ERR", ha="center", va="bottom", color="red", fontweight="bold", fontsize=8)
ax.set_xlabel("Query")
ax.set_ylabel("Number of Datasets Returned")
ax.set_title("Datasets Found per Query: Catalog vs Live")
ax.set_xticks(indices)
ax.set_xticklabels(labels)
ax.legend()
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "04_datasets_found_per_query.png"), dpi=150)
plt.close(fig)

# ── GRAPH 5: Expected ID Hit Rate (stacked bar showing hit/miss) ────────────
fig, ax = plt.subplots(figsize=(14, 6))
cat_hits_int = [1 if h else 0 for h in cat_hit]
live_hits_int = [1 if h else 0 for h in live_hit]
cat_miss_int = [0 if h else 1 for h in cat_hit]
live_miss_int = [0 if h else 1 for h in live_hit]

# Adjust for errors – mark as neither hit nor miss but error
for i in range(n):
    if cat_has_error[i]:
        cat_hits_int[i] = 0
        cat_miss_int[i] = 0
    if live_has_error[i]:
        live_hits_int[i] = 0
        live_miss_int[i] = 0

ax.bar(indices - bar_width/2, cat_hits_int, bar_width, label="Catalog Hit", color="#4CAF50", edgecolor="white")
ax.bar(indices - bar_width/2, cat_miss_int, bar_width, bottom=cat_hits_int, label="Catalog Miss", color="#F44336", edgecolor="white", alpha=0.5)
ax.bar(indices + bar_width/2, live_hits_int, bar_width, label="Live Hit", color="#8BC34A", edgecolor="white")
ax.bar(indices + bar_width/2, live_miss_int, bar_width, bottom=live_hits_int, label="Live Miss", color="#FF5722", edgecolor="white", alpha=0.5)
for i in range(n):
    if cat_has_error[i]:
        ax.text(i - bar_width/2, 0.5, "ERR", ha="center", va="center", color="red", fontweight="bold", fontsize=8)
    if live_has_error[i]:
        ax.text(i + bar_width/2, 0.5, "ERR", ha="center", va="center", color="red", fontweight="bold", fontsize=8)
ax.set_xlabel("Query")
ax.set_ylabel("Hit (1) / Miss (0)")
ax.set_title("Expected Dataset ID Hit: Catalog vs Live")
ax.set_xticks(indices)
ax.set_xticklabels(labels)
ax.set_yticks([0, 1])
ax.set_yticklabels(["Miss", "Hit"])
ax.legend(loc="upper right", fontsize=9)
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "05_expected_id_hit.png"), dpi=150)
plt.close(fig)

# ── GRAPH 6: Summary Statistics (box plots) ─────────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(16, 5))

# Box plot: Tool Calls
bp1 = axes[0].boxplot([cat_tool_calls, live_tool_calls], labels=["Catalog", "Live"],
                       patch_artist=True, widths=0.5)
bp1["boxes"][0].set_facecolor(CAT_COLOR)
bp1["boxes"][1].set_facecolor(LIVE_COLOR)
for b in bp1["boxes"]:
    b.set_alpha(0.7)
axes[0].set_ylabel("Tool Calls")
axes[0].set_title("Tool Calls Distribution")

# Box plot: Trace Steps
bp2 = axes[1].boxplot([cat_steps, live_steps], labels=["Catalog", "Live"],
                       patch_artist=True, widths=0.5)
bp2["boxes"][0].set_facecolor(CAT_COLOR)
bp2["boxes"][1].set_facecolor(LIVE_COLOR)
for b in bp2["boxes"]:
    b.set_alpha(0.7)
axes[1].set_ylabel("Trace Steps")
axes[1].set_title("Trace Steps Distribution")

# Box plot: Execution Time
bp3 = axes[2].boxplot([[t/60 for t in cat_time], [t/60 for t in live_time]],
                       labels=["Catalog", "Live"], patch_artist=True, widths=0.5)
bp3["boxes"][0].set_facecolor(CAT_COLOR)
bp3["boxes"][1].set_facecolor(LIVE_COLOR)
for b in bp3["boxes"]:
    b.set_alpha(0.7)
axes[2].set_ylabel("Execution Time (min)")
axes[2].set_title("Execution Time Distribution")

fig.suptitle("Performance Distributions: Catalog vs Live", fontsize=14, fontweight="bold", y=1.02)
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "06_box_plots.png"), dpi=150, bbox_inches="tight")
plt.close(fig)

# ── GRAPH 7: Scatter – Tool Calls vs Execution Time ─────────────────────────
fig, ax = plt.subplots(figsize=(10, 7))
ax.scatter(cat_tool_calls, [t/60 for t in cat_time], s=100, c=CAT_COLOR, label="Catalog",
           edgecolors=CAT_COLOR2, linewidths=1.5, alpha=0.8, zorder=5)
ax.scatter(live_tool_calls, [t/60 for t in live_time], s=100, c=LIVE_COLOR, label="Live",
           edgecolors=LIVE_COLOR2, linewidths=1.5, alpha=0.8, zorder=5)
# Add query labels
for i in range(n):
    ax.annotate(f"Q{i}", (cat_tool_calls[i], cat_time[i]/60), textcoords="offset points",
                xytext=(5, 5), fontsize=7, color=CAT_COLOR2)
    ax.annotate(f"Q{i}", (live_tool_calls[i], live_time[i]/60), textcoords="offset points",
                xytext=(5, 5), fontsize=7, color=LIVE_COLOR2)
ax.set_xlabel("Number of Tool Calls")
ax.set_ylabel("Execution Time (minutes)")
ax.set_title("Tool Calls vs Execution Time: Catalog vs Live")
ax.legend()
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "07_scatter_calls_vs_time.png"), dpi=150)
plt.close(fig)

# ── GRAPH 8: Efficiency – Datasets found per tool call ──────────────────────
cat_efficiency = []
live_efficiency = []
for i in range(n):
    if cat_tool_calls[i] > 0 and not cat_has_error[i]:
        cat_efficiency.append(cat_dataset_counts[i] / cat_tool_calls[i])
    else:
        cat_efficiency.append(0)
    if live_tool_calls[i] > 0 and not live_has_error[i]:
        live_efficiency.append(live_dataset_counts[i] / live_tool_calls[i])
    else:
        live_efficiency.append(0)

fig, ax = plt.subplots(figsize=(14, 6))
ax.bar(indices - bar_width/2, cat_efficiency, bar_width, label="Catalog", color=CAT_COLOR, edgecolor="white")
ax.bar(indices + bar_width/2, live_efficiency, bar_width, label="Live", color=LIVE_COLOR, edgecolor="white")
ax.set_xlabel("Query")
ax.set_ylabel("Datasets / Tool Call")
ax.set_title("Efficiency: Datasets Found per Tool Call")
ax.set_xticks(indices)
ax.set_xticklabels(labels)
ax.legend()
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "08_efficiency_datasets_per_call.png"), dpi=150)
plt.close(fig)

# ── GRAPH 9: Aggregate Summary Bar ──────────────────────────────────────────
# Exclude error rows for fair comparison
valid_cat = [i for i in range(n) if not cat_has_error[i]]
valid_live = [i for i in range(n) if not live_has_error[i]]

metrics = {
    "Avg Tool Calls": (np.mean(cat_tool_calls), np.mean(live_tool_calls)),
    "Avg Trace Steps": (np.mean(cat_steps), np.mean(live_steps)),
    "Avg Time (min)": (np.mean(cat_time)/60, np.mean(live_time)/60),
    "Avg Datasets Found": (
        np.mean([cat_dataset_counts[i] for i in valid_cat]) if valid_cat else 0,
        np.mean([live_dataset_counts[i] for i in valid_live]) if valid_live else 0,
    ),
    "Hit Rate (%)": (
        100 * sum(cat_hit) / max(len(valid_cat), 1),
        100 * sum(live_hit) / max(len(valid_live), 1),
    ),
}

fig, axes = plt.subplots(1, 5, figsize=(18, 5))
for ax_i, (metric_name, (cat_val, live_val)) in enumerate(metrics.items()):
    ax = axes[ax_i]
    bars = ax.bar(["Catalog", "Live"], [cat_val, live_val], color=[CAT_COLOR, LIVE_COLOR],
                  edgecolor="white", width=0.5)
    ax.set_title(metric_name, fontsize=11)
    # Value labels on bars
    for bar, val in zip(bars, [cat_val, live_val]):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + bar.get_height()*0.02,
                f"{val:.1f}", ha="center", va="bottom", fontsize=10, fontweight="bold")
    ax.set_ylim(0, max(cat_val, live_val) * 1.2)

fig.suptitle("Aggregate Performance Summary: Catalog vs Live", fontsize=14, fontweight="bold", y=1.02)
fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "09_aggregate_summary.png"), dpi=150, bbox_inches="tight")
plt.close(fig)

# ── GRAPH 10: Per-Paper Performance (grouped by paper) ──────────────────────
papers = []
paper_cat_time = {}
paper_live_time = {}
paper_cat_calls = {}
paper_live_calls = {}
for i, row in enumerate(results):
    p = row.get("Paper Short", f"Q{i}")
    if p not in paper_cat_time:
        papers.append(p)
        paper_cat_time[p] = []
        paper_live_time[p] = []
        paper_cat_calls[p] = []
        paper_live_calls[p] = []
    paper_cat_time[p].append(cat_time[i] / 60)
    paper_live_time[p].append(live_time[i] / 60)
    paper_cat_calls[p].append(cat_tool_calls[i])
    paper_live_calls[p].append(live_tool_calls[i])

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10))
p_indices = np.arange(len(papers))

# Average time per paper
cat_p_time = [np.mean(paper_cat_time[p]) for p in papers]
live_p_time = [np.mean(paper_live_time[p]) for p in papers]
ax1.bar(p_indices - bar_width/2, cat_p_time, bar_width, label="Catalog", color=CAT_COLOR, edgecolor="white")
ax1.bar(p_indices + bar_width/2, live_p_time, bar_width, label="Live", color=LIVE_COLOR, edgecolor="white")
ax1.set_ylabel("Avg Execution Time (min)")
ax1.set_title("Average Execution Time by Paper")
ax1.set_xticks(p_indices)
ax1.set_xticklabels(papers, rotation=30, ha="right")
ax1.legend()

# Average tool calls per paper
cat_p_calls = [np.mean(paper_cat_calls[p]) for p in papers]
live_p_calls = [np.mean(paper_live_calls[p]) for p in papers]
ax2.bar(p_indices - bar_width/2, cat_p_calls, bar_width, label="Catalog", color=CAT_COLOR, edgecolor="white")
ax2.bar(p_indices + bar_width/2, live_p_calls, bar_width, label="Live", color=LIVE_COLOR, edgecolor="white")
ax2.set_ylabel("Avg Tool Calls")
ax2.set_title("Average Tool Calls by Paper")
ax2.set_xticks(p_indices)
ax2.set_xticklabels(papers, rotation=30, ha="right")
ax2.legend()

fig.tight_layout()
fig.savefig(os.path.join(OUT_DIR, "10_per_paper_performance.png"), dpi=150)
plt.close(fig)

print(f"✓ Generated 10 graphs in {OUT_DIR}/")
print()

# Print summary table
print("=" * 80)
print("SUMMARY: Catalog vs Live Performance")
print("=" * 80)
print(f"{'Metric':<30} {'Catalog':>12} {'Live':>12} {'Ratio':>10}")
print("-" * 80)
print(f"{'Avg Tool Calls':<30} {np.mean(cat_tool_calls):>12.1f} {np.mean(live_tool_calls):>12.1f} {np.mean(cat_tool_calls)/np.mean(live_tool_calls):>10.1f}x")
print(f"{'Avg Trace Steps':<30} {np.mean(cat_steps):>12.1f} {np.mean(live_steps):>12.1f} {np.mean(cat_steps)/np.mean(live_steps):>10.1f}x")
print(f"{'Avg Exec Time (min)':<30} {np.mean(cat_time)/60:>12.1f} {np.mean(live_time)/60:>12.1f} {np.mean(cat_time)/np.mean(live_time):>10.1f}x")
print(f"{'Total Time (min)':<30} {sum(cat_time)/60:>12.1f} {sum(live_time)/60:>12.1f} {sum(cat_time)/sum(live_time):>10.1f}x")
cat_avg_ds = np.mean([cat_dataset_counts[i] for i in valid_cat]) if valid_cat else 0
live_avg_ds = np.mean([live_dataset_counts[i] for i in valid_live]) if valid_live else 0
print(f"{'Avg Datasets Found':<30} {cat_avg_ds:>12.1f} {live_avg_ds:>12.1f} {cat_avg_ds/max(live_avg_ds,0.01):>10.1f}x")
print(f"{'Expected ID Hit Rate':<30} {100*sum(cat_hit)/max(len(valid_cat),1):>11.0f}% {100*sum(live_hit)/max(len(valid_live),1):>11.0f}%")
print(f"{'Error Count':<30} {sum(cat_has_error):>12d} {sum(live_has_error):>12d}")
print(f"{'Successful Queries':<30} {len(valid_cat):>12d} {len(valid_live):>12d}")
print("=" * 80)

# ── SPREADSHEET: Match rationale ─────────────────────────────────────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MATCH_FILE = os.path.join(BASE, "match_analysis.xlsx")
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Match Analysis"

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
hit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

headers = [
    "Query #", "Paper", "Query", "Expected Identifiers",
    "Live Identifiers", "Catalog Identifiers",
    "Live Match", "Live Match Rationale",
    "Catalog Match", "Catalog Match Rationale",
]
for col, h in enumerate(headers, 1):
    cell = ws_out.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = wrap_align
    cell.border = thin_border

for i, row in enumerate(results):
    r = i + 2  # Excel row (1-indexed, skip header)

    expected = row.get("Expected Identifiers", "")
    live_ids = row.get("Live IDs", "")
    cat_ids = row.get("Catalog IDs", "")

    live_hit, live_rationale = match_expected_ids(expected, live_ids)
    cat_hit_val, cat_rationale = match_expected_ids(expected, cat_ids)

    values = [
        f"Q{i}",
        row.get("Paper Short", ""),
        row.get("Query", ""),
        str(expected or ""),
        str(live_ids or ""),
        str(cat_ids or ""),
        "HIT" if live_hit else "MISS",
        live_rationale,
        "HIT" if cat_hit_val else "MISS",
        cat_rationale,
    ]
    for col, val in enumerate(values, 1):
        cell = ws_out.cell(row=r, column=col, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border

    # Color the match columns
    live_match_cell = ws_out.cell(row=r, column=7)
    live_match_cell.fill = hit_fill if live_hit else miss_fill
    live_match_cell.font = Font(bold=True)

    cat_match_cell = ws_out.cell(row=r, column=9)
    cat_match_cell.fill = hit_fill if cat_hit_val else miss_fill
    cat_match_cell.font = Font(bold=True)

# Column widths
col_widths = [8, 15, 50, 35, 40, 40, 10, 55, 10, 55]
for col, w in enumerate(col_widths, 1):
    ws_out.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = w

# Freeze header row
ws_out.freeze_panes = "A2"

wb_out.save(MATCH_FILE)
print(f"\n✓ Match analysis spreadsheet saved to {MATCH_FILE}")
