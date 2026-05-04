"""Run the unified finder against gold-dataset queries and save results.

Reads queries from ``data/gold_datasets.xlsx``, filters by PDS node,
runs each query through both finder modes (live + catalog), and writes:

    output/<run_name>/
        results.xlsx          — gold sheet with live & catalog result columns
        traces/
            <idx>_live.json   — full message trace for each live run
            <idx>_catalog.json

Usage:
    .venv/bin/python pydantic_code/run_eval.py [--node geo] [--limit 1]
"""

from __future__ import annotations

import argparse
import asyncio
import json
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font

load_dotenv()

try:
    from pydantic_code.finder import FindDatasetOutput, build_finder  # noqa: E402
except ImportError:
    FindDatasetOutput = None  # type: ignore[assignment,misc]
    build_finder = None  # type: ignore[assignment]

# ---------------------------------------------------------------------------
# Gold dataset loader
# ---------------------------------------------------------------------------

_GOLD_FILE = Path(__file__).resolve().parent / "data" / "gold_standard_bundle_collection_dataset_level.xlsx"

# Column indices (0-based) in the gold sheet
_COL_RANK = 0
_COL_PAPER_SHORT = 1
_COL_RATIONALE = 2
_COL_PAPER = 3
_COL_QUERY = 4
_COL_EXPECTED_IDS = 5
_COL_NODE = 6


def load_gold_queries(
    node_filter: str | None = None,
    limit: int | None = None,
) -> list[dict]:
    """Load gold-dataset rows from the Excel file.

    Returns a list of dicts with keys: rank, paper_short, rationale, paper,
    query, expected_ids, node.
    """
    wb = openpyxl.load_workbook(_GOLD_FILE, read_only=True)
    ws = wb["Gold Datasets"]

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        node = str(row[_COL_NODE]).strip().lower() if row[_COL_NODE] else ""
        if node_filter and node != node_filter.lower():
            continue
        rows.append(
            {
                "rank": row[_COL_RANK],
                "paper_short": row[_COL_PAPER_SHORT],
                "rationale": row[_COL_RATIONALE],
                "paper": row[_COL_PAPER],
                "query": row[_COL_QUERY],
                "expected_ids": row[_COL_EXPECTED_IDS],
                "node": row[_COL_NODE],
            }
        )
        if limit and len(rows) >= limit:
            break

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Trace serializer (handles both pydantic models and dataclasses)
# ---------------------------------------------------------------------------


def _serialize(obj):
    if hasattr(obj, "model_dump"):
        return obj.model_dump(mode="json")
    if hasattr(obj, "__dataclass_fields__"):
        return asdict(obj)
    return str(obj)


def _extract_tool_calls(messages) -> list[dict]:
    """Extract a flat list of tool calls from pydantic-ai message history.

    Returns dicts with keys: tool_name, tool_input, tool_output.
    """
    # Build a mapping from tool_call_id → return content
    returns: dict[str, str] = {}
    for msg in messages:
        if msg.kind == "request":
            for part in msg.parts:
                if part.part_kind == "tool-return":
                    returns[part.tool_call_id] = str(part.content)

    # Walk responses and pair each tool call with its return
    calls: list[dict] = []
    for msg in messages:
        if msg.kind == "response":
            for part in msg.parts:
                if part.part_kind == "tool-call":
                    calls.append(
                        {
                            "tool_name": part.tool_name,
                            "tool_input": part.args_as_dict(),
                            "tool_output": returns.get(part.tool_call_id, ""),
                        }
                    )
    return calls


def _is_error_trace(trace: list) -> bool:
    """Return True if *trace* is an error sentinel (not real messages)."""
    return len(trace) == 1 and isinstance(trace[0], dict) and "error" in trace[0]


def _print_tool_calls(tool_calls: list[dict], label: str) -> None:
    """Print tool calls in a clean, readable format."""
    if not tool_calls:
        print(f"  [{label}] No tool calls.")
        return
    for i, tc in enumerate(tool_calls, 1):
        print(f"  [{label}] Tool call {i}:")
        print(f"    tool_name:   {tc['tool_name']}")
        print(f"    tool_input:  {json.dumps(tc['tool_input'], indent=6)}")
        output_str = tc["tool_output"]
        if len(output_str) > 500:
            output_str = output_str[:500] + "... (truncated)"
        print(f"    tool_output: {output_str}")
        print()


# ---------------------------------------------------------------------------
# Single-query runner (returns output + serialized trace)
# ---------------------------------------------------------------------------


async def _run_query(kind: str, query: str, max_retries: int = 2) -> tuple[FindDatasetOutput, list]:
    """Run one query through a finder and return (output, raw_messages).

    Retries up to *max_retries* times when the agent returns zero candidates
    (which usually indicates a spurious early exit or transient failure).
    """
    for attempt in range(1, max_retries + 1):
        agent = build_finder(kind=kind)  # type: ignore[arg-type]
        async with agent:
            result = await agent.run(query)
        output = result.output
        messages = list(result.all_messages())
        if output.candidates or attempt == max_retries:
            return output, messages
        print(f"  [{kind}] Attempt {attempt}: 0 candidates, retrying...")
    return output, messages


# ---------------------------------------------------------------------------
# Format output for Excel cell
# ---------------------------------------------------------------------------


def _format_candidates(output: FindDatasetOutput) -> str:
    """Format candidates into a readable multi-line string for an Excel cell."""
    if not output.candidates:
        return "(no candidates)"
    lines = []
    for i, c in enumerate(output.candidates, 1):
        parts = [f"{i}."]
        if c.dataset_id:
            parts.append(c.dataset_id)
        if c.title:
            parts.append(f'"{c.title}"')
        if c.path:
            parts.append(f"[path={c.path}]")
        if c.node:
            parts.append(f"(node={c.node})")
        if c.pds_version:
            parts.append(f"[{c.pds_version}]")
        lines.append(" ".join(parts))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main eval loop
# ---------------------------------------------------------------------------


async def run_eval(
    node_filter: str | None = "geo",
    limit: int | None = 1,
    mode: str = "both",
    concurrency: int = 3,
) -> None:
    # Load queries
    queries = load_gold_queries(node_filter=node_filter, limit=limit)
    if not queries:
        print(f"No queries found for node={node_filter!r}")
        return

    run_live = mode in ("both", "live")
    run_catalog = mode in ("both", "catalog")
    print(f"Loaded {len(queries)} queries (node={node_filter}, limit={limit}, mode={mode}, concurrency={concurrency})")

    # Create timestamped output directory with separate live/catalog trace dirs
    run_name = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_dir = Path(__file__).resolve().parent / "output" / run_name
    if run_live:
        live_traces_dir = output_dir / "traces" / "live"
        live_traces_dir.mkdir(parents=True, exist_ok=True)
    if run_catalog:
        catalog_traces_dir = output_dir / "traces" / "catalog"
        catalog_traces_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output dir: {output_dir}")

    sem = asyncio.Semaphore(concurrency)

    async def _process_query(idx: int, row: dict) -> dict:
        query = row["query"]

        async with sem:
            print(f"\n[{idx + 1}/{len(queries)}] {query[:80]}...")

            live_output: FindDatasetOutput | None = None
            live_trace: list = []
            catalog_output: FindDatasetOutput | None = None
            catalog_trace: list = []

            # Run live and catalog in parallel when both are enabled
            tasks: dict[str, asyncio.Task] = {}
            if run_live:
                tasks["live"] = asyncio.create_task(_run_query("live", query))
            if run_catalog:
                tasks["catalog"] = asyncio.create_task(_run_query("catalog", query))

            for kind, task in tasks.items():
                try:
                    output, trace = await task
                    print(f"  [{idx + 1}] {kind}: {len(output.candidates)} candidates")
                    if kind == "live":
                        live_output, live_trace = output, trace
                    else:
                        catalog_output, catalog_trace = output, trace
                except Exception as e:
                    print(f"  [{idx + 1}] {kind} FAILED: {e}")
                    error_msg = f"{type(e).__name__}: {e}"
                    if kind == "live":
                        live_trace = [{"error": error_msg}]
                    else:
                        catalog_trace = [{"error": error_msg}]

            # Extract and print tool calls
            if run_live:
                live_tool_calls = _extract_tool_calls(live_trace) if live_trace and not _is_error_trace(live_trace) else []
                _print_tool_calls(live_tool_calls, f"{idx + 1}/live")
            if run_catalog:
                catalog_tool_calls = _extract_tool_calls(catalog_trace) if catalog_trace and not _is_error_trace(catalog_trace) else []
                _print_tool_calls(catalog_tool_calls, f"{idx + 1}/catalog")

            # Save live trace
            if run_live:
                live_trace_data = {
                    "query": query,
                    "row": row,
                    "tool_calls": live_tool_calls,
                    "full_trace": [_serialize(msg) for msg in live_trace],
                }
                (live_traces_dir / f"{idx:03d}_trace.json").write_text(
                    json.dumps(live_trace_data, indent=2, default=str)
                )

            # Save catalog trace
            if run_catalog:
                catalog_trace_data = {
                    "query": query,
                    "row": row,
                    "tool_calls": catalog_tool_calls,
                    "full_trace": [_serialize(msg) for msg in catalog_trace],
                }
                (catalog_traces_dir / f"{idx:03d}_trace.json").write_text(
                    json.dumps(catalog_trace_data, indent=2, default=str)
                )

        return {
            **row,
            "_idx": idx,
            "live_candidates": _format_candidates(live_output) if live_output else ("" if not run_live else "ERROR"),
            "live_summary": live_output.summary if live_output else ("" if not run_live else "ERROR"),
            "live_ids": "\n".join(
                c.dataset_id for c in (live_output.candidates if live_output else []) if c.dataset_id
            ),
            "catalog_candidates": _format_candidates(catalog_output) if catalog_output else ("" if not run_catalog else "ERROR"),
            "catalog_summary": catalog_output.summary if catalog_output else ("" if not run_catalog else "ERROR"),
            "catalog_ids": "\n".join(
                c.dataset_id for c in (catalog_output.candidates if catalog_output else []) if c.dataset_id
            ),
        }

    # Run all queries concurrently (bounded by semaphore)
    raw_results = await asyncio.gather(*[
        _process_query(idx, row) for idx, row in enumerate(queries)
    ])

    # Sort by original index to preserve order in the output
    results = sorted(raw_results, key=lambda r: r.pop("_idx"))

    # Write results Excel
    _write_results_xlsx(output_dir / "results.xlsx", results)
    print(f"\nDone. Results at {output_dir / 'results.xlsx'}")


# ---------------------------------------------------------------------------
# Backfill: re-run only empty catalog traces in an existing output dir
# ---------------------------------------------------------------------------


async def backfill_catalog(output_dir: str | Path) -> None:
    """Scan *output_dir*/traces/catalog/ for empty traces and re-run them serially."""
    output_dir = Path(output_dir).resolve()
    catalog_dir = output_dir / "traces" / "catalog"
    if not catalog_dir.exists():
        print(f"No catalog traces directory at {catalog_dir}")
        return

    # Find trace files with zero tool_calls
    empty_indices: list[tuple[int, dict]] = []
    for trace_file in sorted(catalog_dir.glob("*_trace.json")):
        data = json.loads(trace_file.read_text())
        if not data.get("tool_calls"):
            idx = int(trace_file.stem.split("_")[0])
            empty_indices.append((idx, data["row"]))

    if not empty_indices:
        print("All catalog traces already have tool calls — nothing to backfill.")
        return

    print(f"Found {len(empty_indices)} empty catalog traces: {[i for i, _ in empty_indices]}")
    print("Re-running serially...\n")

    for idx, row in empty_indices:
        query = row["query"]
        print(f"[{idx:03d}] {query[:80]}...")
        try:
            output, trace = await _run_query("catalog", query)
            tool_calls = _extract_tool_calls(trace)
            print(f"  -> {len(output.candidates)} candidates, {len(tool_calls)} tool calls")
            _print_tool_calls(tool_calls, f"{idx:03d}/catalog")
        except Exception as e:
            print(f"  -> FAILED: {e}")
            tool_calls = []
            trace = [{"error": f"{type(e).__name__}: {e}"}]

        trace_data = {
            "query": query,
            "row": row,
            "tool_calls": tool_calls,
            "full_trace": [_serialize(msg) for msg in trace] if not _is_error_trace(trace) else trace,
        }
        (catalog_dir / f"{idx:03d}_trace.json").write_text(
            json.dumps(trace_data, indent=2, default=str)
        )

    print(f"\nBackfill done. Updated traces in {catalog_dir}")


# ---------------------------------------------------------------------------
# Rebuild: reconstruct results.xlsx from existing trace files
# ---------------------------------------------------------------------------


def _extract_output_from_trace(full_trace: list) -> dict | None:
    """Extract the final_result args (candidates + summary) from a full_trace."""
    for msg in reversed(full_trace):
        if not isinstance(msg, dict) or msg.get("kind") != "response":
            continue
        for part in msg.get("parts", []):
            if part.get("part_kind") == "tool-call" and part.get("tool_name") == "final_result":
                args = part.get("args")
                if isinstance(args, str):
                    return json.loads(args)
                return args
    return None


def _format_candidates_from_dict(output: dict | None) -> str:
    """Format candidates dict (from trace) the same way as _format_candidates."""
    if not output or not output.get("candidates"):
        return "(no candidates)"
    lines = []
    for i, c in enumerate(output["candidates"], 1):
        parts = [f"{i}."]
        if c.get("dataset_id"):
            parts.append(c["dataset_id"])
        if c.get("title"):
            parts.append(f'"{c["title"]}"')
        if c.get("path"):
            parts.append(f"[path={c['path']}]")
        if c.get("node"):
            parts.append(f"(node={c['node']})")
        if c.get("pds_version"):
            parts.append(f"[{c['pds_version']}]")
        lines.append(" ".join(parts))
    return "\n".join(lines)


def rebuild_results(output_dir: str | Path) -> None:
    """Reconstruct results.xlsx from trace files in an existing output directory."""
    output_dir = Path(output_dir).resolve()
    live_dir = output_dir / "traces" / "live"
    catalog_dir = output_dir / "traces" / "catalog"

    has_live = live_dir.exists()
    has_catalog = catalog_dir.exists()
    if not has_live and not has_catalog:
        print(f"No trace directories found in {output_dir}")
        return

    # Collect all trace indices
    indices: set[int] = set()
    for d in [live_dir, catalog_dir]:
        if d.exists():
            for f in d.glob("*_trace.json"):
                indices.add(int(f.stem.split("_")[0]))

    results = []
    for idx in sorted(indices):
        row: dict = {}
        live_output: dict | None = None
        catalog_output: dict | None = None

        # Read live trace
        if has_live:
            live_file = live_dir / f"{idx:03d}_trace.json"
            if live_file.exists():
                data = json.loads(live_file.read_text())
                row = data.get("row", row)
                live_output = _extract_output_from_trace(data.get("full_trace", []))

        # Read catalog trace
        if has_catalog:
            catalog_file = catalog_dir / f"{idx:03d}_trace.json"
            if catalog_file.exists():
                data = json.loads(catalog_file.read_text())
                row = row or data.get("row", row)
                catalog_output = _extract_output_from_trace(data.get("full_trace", []))

        results.append({
            **row,
            "live_ids": "\n".join(
                c["dataset_id"] for c in (live_output or {}).get("candidates", []) if c.get("dataset_id")
            ),
            "live_candidates": _format_candidates_from_dict(live_output),
            "live_summary": (live_output or {}).get("summary", ""),
            "catalog_ids": "\n".join(
                c["dataset_id"] for c in (catalog_output or {}).get("candidates", []) if c.get("dataset_id")
            ),
            "catalog_candidates": _format_candidates_from_dict(catalog_output),
            "catalog_summary": (catalog_output or {}).get("summary", ""),
        })

    xlsx_path = output_dir / "results.xlsx"
    _write_results_xlsx(xlsx_path, results)
    print(f"Rebuilt {len(results)} rows -> {xlsx_path}")


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

_RESULT_HEADERS = [
    "Geoscience Rank",
    "Paper Short",
    "Geoscience Rationale",
    "Paper",
    "Query",
    "Expected Identifiers",
    "PDS Node",
    "Live IDs",
    "Live Candidates",
    "Live Summary",
    "Catalog IDs",
    "Catalog Candidates",
    "Catalog Summary",
]

_RESULT_KEYS = [
    "rank",
    "paper_short",
    "rationale",
    "paper",
    "query",
    "expected_ids",
    "node",
    "live_ids",
    "live_candidates",
    "live_summary",
    "catalog_ids",
    "catalog_candidates",
    "catalog_summary",
]


def _write_results_xlsx(path: Path, results: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eval Results"

    # Header row
    header_font = Font(bold=True)
    for col, header in enumerate(_RESULT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Data rows
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, key in enumerate(_RESULT_KEYS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
            cell.alignment = wrap

    # Auto-width for key columns (approximate)
    for col_idx, header in enumerate(_RESULT_HEADERS, 1):
        if "Candidates" in header or "Summary" in header:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 60
        elif "IDs" in header or "Identifiers" in header:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 40
        elif header == "Query":
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 50
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    wb.save(path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Run PDS finder eval against gold datasets")
    parser.add_argument("--node", default="geo", help="PDS node to filter by (default: geo)")
    parser.add_argument("--limit", type=int, default=1, help="Max queries to run (default: 1, 0 = no limit)")
    parser.add_argument(
        "--mode",
        choices=["both", "live", "catalog"],
        default="both",
        help="Which finder mode(s) to run (default: both)",
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=3,
        help="Max queries to run concurrently (default: 3)",
    )
    parser.add_argument(
        "--backfill",
        type=str,
        default=None,
        metavar="OUTPUT_DIR",
        help="Re-run only empty catalog traces in an existing output directory",
    )
    parser.add_argument(
        "--rebuild",
        type=str,
        default=None,
        metavar="OUTPUT_DIR",
        help="Rebuild results.xlsx from existing trace files in an output directory",
    )
    args = parser.parse_args()

    if args.rebuild:
        rebuild_results(args.rebuild)
    elif args.backfill:
        asyncio.run(backfill_catalog(args.backfill))
    else:
        asyncio.run(run_eval(node_filter=args.node, limit=args.limit or None, mode=args.mode, concurrency=args.concurrency))


if __name__ == "__main__":
    main()
